"""
AARO ERP - PDKS Export Yöneticisi
"""

import csv
from tkinter import filedialog, messagebox

class ExportManager:
    def __init__(self):
        pass
    
    def export_data(self, parent, users_data, attendance_data):
        """Verileri dışa aktar"""
        if not users_data and not attendance_data:
            messagebox.showinfo("Bilgi", "Aktarılacak veri yok.")
            return
        
        # Dosya kaydetme dialog'u
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel dosyaları", "*.xlsx"), ("CSV dosyaları", "*.csv"), ("Tüm dosyalar", "*.*")],
            title="Dosya Kaydet"
        )
        
        if filename:
            try:
                if filename.endswith('.xlsx'):
                    self.export_to_excel(filename, users_data, attendance_data)
                else:
                    self.export_to_csv(filename, users_data, attendance_data)
                messagebox.showinfo("Başarılı", f"Veriler başarıyla aktarıldı: {filename}")
            except Exception as e:
                messagebox.showerror("Hata", f"Dosya kaydedilemedi: {str(e)}")
    
    def export_to_excel(self, filename, users_data, attendance_data):
        """Excel dosyasına aktar"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Hata", "openpyxl kütüphanesi yüklü değil. 'pip install openpyxl' komutu ile yükleyin.")
            return
        
        wb = openpyxl.Workbook()
        
        # Kullanıcılar sayfası
        if users_data:
            ws_users = wb.active
            ws_users.title = "Kullanıcılar"
            
            # Başlık satırı
            headers_users = ["Kullanıcı ID", "Ad Soyad", "Yetki", "Cihaz"]
            for col, header in enumerate(headers_users, 1):
                cell = ws_users.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Kullanıcı verileri
            for row, user in enumerate(users_data, 2):
                ws_users.cell(row=row, column=1, value=user['uid'])
                ws_users.cell(row=row, column=2, value=user['name'])
                ws_users.cell(row=row, column=3, value=user['privilege_text'])
                ws_users.cell(row=row, column=4, value=user['device_name'])
        
        # Giriş-Çıkış kayıtları sayfası
        if attendance_data:
            if users_data:
                ws_attendance = wb.create_sheet("Giriş-Çıkış Kayıtları")
            else:
                ws_attendance = wb.active
                ws_attendance.title = "Giriş-Çıkış Kayıtları"
            
            # Başlık satırı
            headers_attendance = ["Kullanıcı ID", "Ad Soyad", "Tarih/Saat", "Durum", "Cihaz"]
            for col, header in enumerate(headers_attendance, 1):
                cell = ws_attendance.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Attendance verileri
            for row, att in enumerate(attendance_data, 2):
                ws_attendance.cell(row=row, column=1, value=att['uid'])
                ws_attendance.cell(row=row, column=2, value=att['user_name'])
                ws_attendance.cell(row=row, column=3, value=att['timestamp_str'])
                ws_attendance.cell(row=row, column=4, value=att['status_text'])
                ws_attendance.cell(row=row, column=5, value=att['device_name'])
        
        # Sütun genişliklerini ayarla
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
    
    def export_to_csv(self, filename, users_data, attendance_data):
        """CSV dosyasına aktar"""
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Kullanıcılar bölümü
            if users_data:
                writer.writerow(["=== KULLANICILAR ==="])
                writer.writerow(["Kullanıcı ID", "Ad Soyad", "Yetki", "Cihaz"])
                
                for user in users_data:
                    writer.writerow([user['uid'], user['name'], user['privilege_text'], user['device_name']])
                
                writer.writerow([])  # Boş satır
            
            # Giriş-Çıkış kayıtları bölümü
            if attendance_data:
                writer.writerow(["=== GİRİŞ-ÇIKIŞ KAYITLARI ==="])
                writer.writerow(["Kullanıcı ID", "Ad Soyad", "Tarih/Saat", "Durum", "Cihaz"])
                
                for att in attendance_data:
                    writer.writerow([att['uid'], att['user_name'], att['timestamp_str'], att['status_text'], att['device_name']])
